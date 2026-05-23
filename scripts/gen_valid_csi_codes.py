"""Read the canonical Buildertrend CSI cost code list and emit a JS object literal."""
import openpyxl
import re
import json
import os

SRC = r"C:\Users\MattFarrier\Downloads\4.17.26Buildertrend_CSI_Cost_Codes.xlsx"
OUT = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\scripts\valid_csi_codes_js.txt"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active

# Inspect first few rows to find header
print("First 5 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
    print(i, row)

codes = {}
# Try columns A and B; assume row 1 is header
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row: continue
    a = row[0] if len(row) > 0 else None
    b = row[1] if len(row) > 1 else None
    if a is None: continue
    a_str = str(a).strip()
    b_str = str(b).strip() if b is not None else ""
    if not a_str: continue
    # Expect format like "1.101 Project Administration and General Office Labor" OR just "1.101" with label in B
    # Match leading numeric prefix
    m = re.match(r"^(\d+)\.(\d+)\s*(.*)$", a_str)
    if not m:
        continue
    div = str(int(m.group(1)))  # strip leading zeros
    sub = m.group(2)
    rest = m.group(3).strip()
    prefix = f"{div}.{sub}"
    # Canonical label: full string with leading-zero-stripped prefix
    if rest:
        label = f"{prefix} {rest}"
    elif b_str:
        label = f"{prefix} {b_str}"
    else:
        label = prefix
    codes[prefix] = label

print(f"\nTotal codes: {len(codes)}")

# Emit JS
lines = ["const VALID_CSI_CODES = {"]
# Sort by (division int, subcode int)
def sort_key(k):
    a, b = k.split(".")
    return (int(a), int(b))
keys = sorted(codes.keys(), key=sort_key)
for i, k in enumerate(keys):
    label = codes[k].replace('"', '\\"')
    comma = "," if i < len(keys) - 1 else ""
    lines.append(f'  "{k}": "{label}"{comma}')
lines.append("};")
out = "\n".join(lines)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write(out)

print(f"Wrote {OUT}")
print(f"Key count: {len(keys)}")
print("First 3:", keys[:3])
print("Last 3:", keys[-3:])
