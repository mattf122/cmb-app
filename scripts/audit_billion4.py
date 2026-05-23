"""Full issue table + deeper look at the worst offenders."""
import re
from openpyxl import load_workbook

ESTIMATE = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Estimate_2026-04-24.xlsx"
CSI_REF = r"C:\Users\MattFarrier\Downloads\4.17.26Buildertrend_CSI_Cost_Codes.xlsx"

wb = load_workbook(CSI_REF, data_only=True)
ws = wb.active
csi_codes = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    code = str(row[0]).strip()
    csi_codes[code] = str(row[1]).strip() if row[1] else ""

# Build set of valid numeric prefixes
valid_prefixes = set()
for c in csi_codes:
    m = re.match(r"^(\d+\.\d+)", c)
    if m:
        valid_prefixes.add(m.group(1))

wb2 = load_workbook(ESTIMATE, data_only=True)
ws2 = wb2.active
rows = []
for r in range(2, ws2.max_row + 1):
    row = [ws2.cell(row=r, column=c).value for c in range(1, 16)]
    if all(v is None or v == "" for v in row):
        continue
    rows.append((r, row))

SUFFIX_TO_TYPE = {"01": "Labor", "02": "Material", "03": "Subcontractor"}
issues = []
for r, row in rows:
    cat, code, title, _, _, _, _, ctype = row[:8]
    code_s = str(code or "").strip()
    first_tok = code_s.split()[0] if code_s else ""

    problems = []
    # Exact CSI membership
    if code_s not in csi_codes:
        # Is it same as CSI entry with leading zero stripped from prefix?
        # CSI uses "1.101 Name" not "01.101 Name"
        stripped = first_tok.lstrip("0") if first_tok else ""
        candidates = [c for c in csi_codes if c.startswith(stripped + " ")]
        if first_tok and first_tok != stripped and candidates:
            problems.append(f"leading-zero: '{first_tok}' should be '{stripped}'")
        elif stripped in valid_prefixes or first_tok in valid_prefixes:
            # prefix valid but label text doesn't match CSI exactly
            problems.append(f"prefix ok but full value != any CSI key")
        else:
            problems.append(f"prefix '{first_tok}' not a valid CSI prefix")

    # suffix check
    m = re.match(r"^\d+\.(\d+)", first_tok)
    suffix = m.group(1)[-2:] if m and len(m.group(1)) >= 2 else ""
    if suffix in SUFFIX_TO_TYPE:
        if ctype != SUFFIX_TO_TYPE[suffix]:
            problems.append(f"suffix {suffix} expects {SUFFIX_TO_TYPE[suffix]}, got {ctype!r}")
    elif len(suffix) == 4:  # e.g. 21.0103 -> picks '03' but got 0103
        last2 = suffix[-2:]
        if last2 in SUFFIX_TO_TYPE and ctype != SUFFIX_TO_TYPE[last2]:
            problems.append(f"odd 4-digit tail, last2={last2}")

    # division check
    cm = re.match(r"^(\d+)\.", first_tok)
    code_div = cm.group(1).lstrip("0") if cm else ""
    cm2 = re.match(r"^(\d+)", str(cat or ""))
    cat_div = cm2.group(1).lstrip("0") if cm2 else ""
    if code_div and cat_div and code_div != cat_div:
        problems.append(f"code div {code_div} != category div {cat_div}")

    if problems:
        issues.append((r, str(cat or ""), code_s, str(title or ""), str(ctype or ""), "; ".join(problems)))

print(f"Total issues: {len(issues)} of {len(rows)}\n")
print(f"{'Row':<4} {'Category':<26} {'Code':<55} {'Title':<35} {'Type':<13} Problem")
for iss in issues:
    r, cat, code, title, ct, prob = iss
    print(f"{r:<4} {cat[:25]:<26} {code[:54]:<55} {title[:34]:<35} {ct[:12]:<13} {prob}")

# The 21.xxxx codes — let's see if CSI has 4-digit codes
print("\nCSI codes starting with 21.:")
for c in sorted(csi_codes):
    if c.startswith("21."):
        print(f"  {c}")
