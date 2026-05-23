from openpyxl import load_workbook
F = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\sample_BT_v2.xlsx"
wb = load_workbook(F, data_only=False)
ws = wb.active
print(f"Sheet name: {ws.title!r}")
print(f"Max row: {ws.max_row}, max col: {ws.max_column}")
print(f"Merged cells: {list(ws.merged_cells.ranges)}")
print(f"Freeze panes: {ws.freeze_panes}")

# Column widths
print("\nColumn widths:")
for col in "ABCDEFGHIJKLMNO":
    print(f"  {col}: {ws.column_dimensions[col].width}")

# Header check
print("\nRow 1 (headers):")
for c in ws[1]:
    print(f"  {c.coordinate}: {c.value!r} font={c.font.name} sz={c.font.size} bold={c.font.bold} fill={c.fill.fgColor.rgb if c.fill.fgColor else None}")

# Sample data row (row 2) formats
print("\nRow 2 formats:")
for c in ws[2]:
    print(f"  {c.coordinate}: val={c.value!r} fmt={c.number_format} font={c.font.name} sz={c.font.size} bold={c.font.bold}")

# First 5 data rows
print("\n=== First 5 data rows ===")
for r in range(2, 7):
    vals = [ws.cell(row=r, column=ci).value for ci in range(1,16)]
    print(r, vals)

# Last 3 data rows
print("\n=== Last 3 data rows ===")
for r in range(ws.max_row-2, ws.max_row+1):
    vals = [ws.cell(row=r, column=ci).value for ci in range(1,16)]
    print(r, vals)

# Check for any cell with fill or border in data body
issues = 0
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for c in row:
        if c.fill and c.fill.fgColor and c.fill.fgColor.rgb and c.fill.fgColor.rgb not in ("00000000",None):
            if c.fill.patternType:
                issues += 1
print(f"\nCells with fill patternType set: {issues}")

# Check formulas present in J,M,N,O for every data row
missing = []
for r in range(2, ws.max_row+1):
    for col, pref in [(10,"=E"),(13,"=J"),(14,"=IF(M"),(15,"=M")]:
        v = ws.cell(row=r, column=col).value
        if not (isinstance(v, str) and v.startswith(pref)):
            missing.append((r,col,v))
print(f"Rows missing expected formulas: {len(missing)}")
if missing[:3]:
    print("  sample:", missing[:3])

# Scan for subtotal/total/blank rows
bad = []
for r in range(2, ws.max_row+1):
    vals = [ws.cell(row=r, column=ci).value for ci in range(1,16)]
    if all(v is None or v == "" for v in vals):
        bad.append((r,"BLANK"))
        continue
    txt = " ".join(str(v) for v in vals if v is not None).lower()
    if "subtotal" in txt or "grand total" in txt or txt.strip() == "total":
        bad.append((r,"TOTAL",txt[:80]))
print(f"Body bad rows (blank/subtotal/total): {len(bad)}")
if bad:
    print("  ", bad)

print(f"\nFinal data row count: {ws.max_row - 1}")
