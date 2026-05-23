"""Follow-up: inspect Excel cell types (formula vs value) and proposal structure."""
import re
from openpyxl import load_workbook

ESTIMATE = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Estimate_2026-04-24.xlsx"
PROPOSAL = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Proposal_2026-04-24.doc"

# 1. Excel: inspect formulas stored
print("=== Excel - inspect with formulas (data_only=False) ===")
wb = load_workbook(ESTIMATE, data_only=False)
ws = wb.active
print(f"Row 2 raw (col by col):")
for c in range(1, 16):
    v = ws.cell(row=2, column=c).value
    print(f"  col {c}: {v!r}")

print("\nRow 5 raw:")
for c in range(1, 16):
    v = ws.cell(row=5, column=c).value
    print(f"  col {c}: {v!r}")

# Also load with data_only and recompute totals manually from unit_cost * qty
print("\n=== Manual compute of Builder Cost = Qty * UnitCost for all rows ===")
wb2 = load_workbook(ESTIMATE, data_only=True)
ws2 = wb2.active
rows = []
for r in range(2, ws2.max_row + 1):
    row = [ws2.cell(row=r, column=c).value for c in range(1, 16)]
    if all(v is None or v == "" for v in row):
        continue
    rows.append((r, row))

total_bc_manual = 0.0
total_cp_manual = 0.0
per_cat = {}
for r, row in rows:
    qty = row[4]
    uc = row[6]
    markup = row[10] or 0
    cat = row[0]
    try:
        bc = float(qty) * float(uc) if qty is not None and uc is not None else 0
    except (TypeError, ValueError):
        bc = 0
    try:
        cp = bc * (1 + float(markup)/100)
    except (TypeError, ValueError):
        cp = bc
    total_bc_manual += bc
    total_cp_manual += cp
    per_cat.setdefault(cat, {"bc":0.0,"cp":0.0,"n":0})
    per_cat[cat]["bc"] += bc
    per_cat[cat]["cp"] += cp
    per_cat[cat]["n"] += 1

print(f"Manual BC total: ${total_bc_manual:,.2f}")
print(f"Manual CP total: ${total_cp_manual:,.2f}")
print("\nPer-category:")
for cat, v in sorted(per_cat.items(), key=lambda x: str(x[0])):
    print(f"  {str(cat):<40} n={v['n']:<4} BC=${v['bc']:>13,.2f}  CP=${v['cp']:>13,.2f}")

# 2. Proposal: dump structure
print("\n=== Proposal raw bytes peek ===")
with open(PROPOSAL, "rb") as f:
    raw = f.read()
print(f"Total bytes: {len(raw)}")
print(f"First 500 bytes decoded utf-8:\n{raw[:500].decode('utf-8', errors='replace')}")
print(f"\nLast 1200 bytes:\n{raw[-1200:].decode('utf-8', errors='replace')}")

text = raw.decode("utf-8", errors="replace")
# Search any $ signs
dollar_count = text.count("$")
print(f"\nTotal '$' chars: {dollar_count}")
# Print all segments with $
for m in re.finditer(r".{0,40}\$[^<\n]{0,80}", text):
    s = m.group(0)
    if "$" in s:
        print(f"  ...{s[:120]}")
