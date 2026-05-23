"""Load the computed copy and print first 5 and last 3 data rows with
the fields that matter for formula verification."""
from openpyxl import load_workbook

PATH = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\sample_BT_import_computed.xlsx"

wb = load_workbook(PATH, data_only=True)
ws = wb.active

HEADERS = [c.value for c in ws[1]]
max_row = ws.max_row
print(f"Sheet: {ws.title}  Rows: {max_row}  Cols: {ws.max_column}")
print(f"Data rows: {max_row - 1}")
print()

def dump(r):
    row = {HEADERS[i]: ws.cell(row=r, column=i+1).value for i in range(len(HEADERS))}
    keys = ["Category", "Cost Code", "Title", "Quantity", "Unit", "Unit Cost",
            "Cost Type", "Builder Cost", "Markup", "Unit Price", "Client Price",
            "Margin", "Profit"]
    line = f"Row {r}: " + " | ".join(f"{k}={row[k]}" for k in keys)
    print(line)

print("=== First 5 data rows ===")
for r in range(2, 7):
    dump(r)

print()
print("=== Last 3 data rows ===")
for r in range(max_row - 2, max_row + 1):
    dump(r)
