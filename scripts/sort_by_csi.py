import openpyxl
from copy import copy

PATH = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\sample_BT_v2.xlsx"

wb = openpyxl.load_workbook(PATH)
ws = wb.active

max_row = ws.max_row
max_col = ws.max_column
print(f"Before: max_row={max_row}, max_col={max_col}")

# Read data rows (2..max_row), columns A..O (1..15)
# For columns J, M, N, O we skip reading (they're formulas we'll regenerate)
FORMULA_COLS = {10, 13, 14, 15}  # J, M, N, O

rows_data = []
row_styles = []  # per-row list of cell styles for each column

for r in range(2, max_row + 1):
    row_vals = {}
    row_style = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        if c in FORMULA_COLS:
            row_vals[c] = None  # will regenerate
        else:
            row_vals[c] = cell.value
        # capture style
        row_style[c] = {
            'font': copy(cell.font),
            'fill': copy(cell.fill),
            'border': copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
            'protection': copy(cell.protection),
        }
    rows_data.append(row_vals)
    row_styles.append(row_style)

print(f"Read {len(rows_data)} data rows")

# Stable sort by leading number of Category (column A = 1)
def sort_key(row):
    cat = row[1]
    try:
        return int(str(cat).split(' ')[0])
    except Exception:
        return 9999

indexed = list(range(len(rows_data)))
indexed.sort(key=lambda i: sort_key(rows_data[i]))

sorted_data = [rows_data[i] for i in indexed]
sorted_styles = [row_styles[i] for i in indexed]

# Clear existing data rows (values and formulas); keep header row
for r in range(2, max_row + 1):
    for c in range(1, max_col + 1):
        ws.cell(row=r, column=c).value = None

# Write sorted data back starting at row 2
for idx, (rv, rs) in enumerate(zip(sorted_data, sorted_styles)):
    r = idx + 2
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        if c == 10:   # J
            cell.value = f"=E{r}*G{r}"
        elif c == 13: # M
            cell.value = f"=J{r}*(1+K{r}/100)"
        elif c == 14: # N
            cell.value = f"=IF(M{r}=0,0,O{r}/M{r})"
        elif c == 15: # O
            cell.value = f"=M{r}-J{r}"
        else:
            cell.value = rv[c]
        # reapply style
        style = rs[c]
        cell.font = style['font']
        cell.fill = style['fill']
        cell.border = style['border']
        cell.alignment = style['alignment']
        cell.number_format = style['number_format']
        cell.protection = style['protection']

wb.save(PATH)
print("Saved.")

# Verify
wb2 = openpyxl.load_workbook(PATH)
ws2 = wb2.active
print(f"\nAfter: max_row={ws2.max_row}")
print("\nCategory column top-to-bottom:")
for r in range(2, ws2.max_row + 1):
    print(f"  row {r}: {ws2.cell(row=r, column=1).value}")

print("\nFirst 5 data rows (Category | Title):")
for r in range(2, 7):
    print(f"  row {r}: {ws2.cell(row=r, column=1).value} | {ws2.cell(row=r, column=2).value}")

print("\nLast 3 data rows (Category | Title):")
for r in range(ws2.max_row - 2, ws2.max_row + 1):
    print(f"  row {r}: {ws2.cell(row=r, column=1).value} | {ws2.cell(row=r, column=2).value}")

# Formula spot-check - load with data_only=False (default)
print("\nFormula spot-check:")
for r in [2, 10, ws2.max_row]:
    print(f"  row {r}: J={ws2.cell(row=r,column=10).value} | M={ws2.cell(row=r,column=13).value} | N={ws2.cell(row=r,column=14).value} | O={ws2.cell(row=r,column=15).value}")

print(f"\nTotal row count: header + {ws2.max_row - 1} data rows = {ws2.max_row} rows")
