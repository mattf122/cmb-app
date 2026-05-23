"""Build sample_BT_v2.xlsx matching the template format exactly."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

OUT = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\sample_BT_v2.xlsx"

HEADERS = ["Category","Cost Code","Title","Description","Quantity","Unit","Unit Cost",
           "Cost Type","Marked As","Builder Cost","Markup","Markup Type","Client Price","Margin","Profit"]

COL_WIDTHS = {"A":42,"B":48,"C":32,"D":42,"E":11,"F":8,"G":13,"H":15,"I":13,
              "J":14,"K":10,"L":13,"M":14,"N":11,"O":13}

# Category constants
CAT_DEMO = "02 Existing Conditions"
CAT_WOOD = "06 Woods, Plastics, and Composites"
CAT_FURN = "12 Furnishings"
CAT_FIN  = "09 Finishes"
CAT_PLMB = "22 Plumbing"
CAT_ELEC = "26 Electrical"
CAT_EQUIP= "11 Equipment"
CAT_GC   = "01 General Requirements"
CAT_PROC = "00 Procurement and Contracting Requirements"

# Each row: (Category, CostCode, Title, Description, Qty, Unit, UnitCost, CostType, MarkedAs)
rows = []

# --- DEMOLITION ---
# 1 Cabinet Demo
rows.append((CAT_DEMO,"2.201","Cabinet Demo Labor","Remove existing uppers and lowers",8,"HR",85,"Labor","Estimate"))
rows.append((CAT_DEMO,"2.202","Cabinet Demo Disposal","Haul-away & dump fees",22,"LF",25,"Material","Estimate"))
# 2 Countertop & Backsplash
rows.append((CAT_DEMO,"2.201","Countertop Removal Labor","Carefully demo tops and tile",4,"HR",85,"Labor","Estimate"))
rows.append((CAT_DEMO,"2.202","Countertop Disposal","Haul-away & dump fees",38,"SF",8,"Material","Estimate"))
# 3 Load-Bearing Wall
rows.append((CAT_DEMO,"2.201","Wall Demo w/ Temp Shoring","Remove wall, install temp support",16,"HR",100,"Labor","Estimate"))
rows.append((CAT_DEMO,"2.202","Shoring + Disposal","Temp shoring lumber & debris removal",1,"EA",1200,"Material","Estimate"))
# 4 Flooring Removal
rows.append((CAT_DEMO,"2.201","Floor Demo Labor","Remove existing flooring to subfloor",6,"HR",85,"Labor","Estimate"))
rows.append((CAT_DEMO,"2.202","Floor Demo Disposal","Haul-away & dump fees",280,"SF",2,"Material","Estimate"))

# --- ROUGH CARPENTRY ---
# 5 Blocking & Backing
rows.append((CAT_WOOD,"6.101","Blocking Labor","Backing for cabinets and accessories",6,"HR",85,"Labor","Estimate"))
rows.append((CAT_WOOD,"6.102","Blocking Material","Lumber for blocking/backing",1,"LS",280,"Material","Estimate"))
# 6 Structural Beam
rows.append((CAT_WOOD,"6.101","Beam Install Labor","Set LVL beam per engineer's detail",12,"HR",100,"Labor","Estimate"))
rows.append((CAT_WOOD,"6.102","LVL Beam & Fasteners","LVL beam with fasteners and hangers",1,"EA",1800,"Material","Estimate"))
# 7 Subfloor Patching
rows.append((CAT_WOOD,"6.101","Subfloor Repair Labor","Patch subfloor as needed",6,"HR",85,"Labor","Estimate"))
rows.append((CAT_WOOD,"6.102","Subfloor Patch Material",'3/4" AdvanTech patches',1,"LS",220,"Material","Estimate"))

# --- CABINETRY & COUNTERTOPS ---
# 8 White Shaker
rows.append((CAT_FURN,"12.103","Kitchen Cabinets – Shaker White","Cabinetry supply & install, soft-close",30,"LF",650,"Subcontractor","Estimate"))
# 9 Island Base
rows.append((CAT_FURN,"12.103","Island Cabinets","Base cabinets for island w/ seating side",8,"LF",700,"Subcontractor","Estimate"))
# 10 Quartz
rows.append((CAT_FURN,"12.103","Quartz Countertops",'Supply & install, perimeter + 4" splash',38,"SF",150,"Subcontractor","Allowance"))
# 11 Waterfall
rows.append((CAT_FURN,"12.103","Waterfall Edge Island Top","Fabrication and install, waterfall sides",24,"SF",220,"Subcontractor","Allowance"))

# --- FINISHES ---
# 12 Subway Tile
rows.append((CAT_FIN,"9.303","Tile Backsplash","Install subway tile, grout & seal",38,"SF",32,"Subcontractor","Allowance"))
# 13 LVP
rows.append((CAT_FIN,"9.503","LVP Flooring","Supply & install LVP plank flooring",280,"SF",11,"Subcontractor","Allowance"))
# 14 Paint
rows.append((CAT_FIN,"9.803","Interior Paint","Prime + 2 coats, walls and ceiling",1,"LS",1400,"Subcontractor","Estimate"))

# --- PLUMBING ---
# 15 Sink
rows.append((CAT_PLMB,"22.203","Kitchen Plumbing","Relocate sink rough-in, install sink/faucet",1,"LS",2400,"Subcontractor","Estimate"))
# 16 DW
rows.append((CAT_PLMB,"22.203","DW Supply & Drain","Connect water and drain for DW",1,"LS",450,"Subcontractor","Estimate"))

# --- ELECTRICAL ---
# 17 Cans
rows.append((CAT_ELEC,"26.303","Recessed Cans",'6" LED cans, supply & install',8,"EA",175,"Subcontractor","Allowance"))
# 18 Under-Cab
rows.append((CAT_ELEC,"26.303","Under-Cab Lighting","LED strip lighting, hardwired",22,"LF",55,"Subcontractor","Estimate"))
# 19 Pendants
rows.append((CAT_ELEC,"26.303","Island Pendants","3 pendants, supply & install",3,"EA",420,"Subcontractor","Allowance"))
# 20 Devices
rows.append((CAT_ELEC,"26.203","Kitchen Device Work","Relocate/add outlets & switches, GFCI at counters",1,"LS",850,"Subcontractor","Estimate"))

# --- APPLIANCES ---
# 21
rows.append((CAT_EQUIP,"11.202","Appliance Allowance","Range, hood, DW, fridge, microwave",1,"LOT",12000,"Material","Allowance"))

# --- GENERAL CONDITIONS ---
# 22 PM
rows.append((CAT_GC,"1.201","Project Management","PM time — coordination, scheduling, owner comm",40,"HR",130,"Labor","Estimate"))
# 23 Super
rows.append((CAT_GC,"1.201","Superintendent","On-site supervision",20,"HR",100,"Labor","Estimate"))
# 24 Dumpster
rows.append((CAT_GC,"1.901","Dumpster Labor","Loading/hauling",2,"HR",85,"Labor","Estimate"))
rows.append((CAT_GC,"1.902","Dumpster Rental","20-yd container + 2 swaps",1,"LS",950,"Material","Estimate"))
# 25 Cleanup
rows.append((CAT_GC,"1.901","Daily Cleanup","Daily broom-clean and final clean",16,"HR",85,"Labor","Estimate"))
rows.append((CAT_GC,"1.902","Cleaning Supplies","Consumables for cleanup",1,"LS",150,"Material","Estimate"))
# 26 Protection
rows.append((CAT_GC,"1.301","Protection Labor","Install/remove protection",4,"HR",85,"Labor","Estimate"))
rows.append((CAT_GC,"1.302","Protection Material","Ram Board, zip walls, plastic",1,"LS",320,"Material","Estimate"))
# 27 GL
rows.append((CAT_GC,"1.701","GL Insurance Allocation","Prorated GL insurance for job duration",1,"LS",650,"Labor","Estimate"))

# --- PERMITS / DESIGN ---
# 28 Permit
rows.append((CAT_PROC,"0.101","Building Permit","City/county permit fees",1,"LS",650,"Labor","Allowance"))
# 29 Engineer
rows.append((CAT_PROC,"0.203","Engineering","Structural calc for beam",1,"LS",850,"Subcontractor","Allowance"))

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# Column widths
for col, w in COL_WIDTHS.items():
    ws.column_dimensions[col].width = w

# Header row
bold_font = Font(name="Calibri", size=11, bold=True)
reg_font  = Font(name="Calibri", size=11, bold=False)

for i, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = bold_font

# Data rows
for idx, row in enumerate(rows, start=2):
    cat, code, title, desc, qty, unit, unit_cost, cost_type, marked = row
    ws.cell(row=idx, column=1, value=cat).font = reg_font
    ws.cell(row=idx, column=2, value=code).font = reg_font
    ws.cell(row=idx, column=3, value=title).font = reg_font
    ws.cell(row=idx, column=4, value=desc).font = reg_font
    e = ws.cell(row=idx, column=5, value=qty); e.font = reg_font; e.number_format = "#,##0.00"
    ws.cell(row=idx, column=6, value=unit).font = reg_font
    g = ws.cell(row=idx, column=7, value=unit_cost); g.font = reg_font; g.number_format = "$#,##0.00"
    ws.cell(row=idx, column=8, value=cost_type).font = reg_font
    ws.cell(row=idx, column=9, value=marked).font = reg_font
    j = ws.cell(row=idx, column=10, value=f"=E{idx}*G{idx}"); j.font = reg_font; j.number_format = "$#,##0.00"
    k = ws.cell(row=idx, column=11, value=20); k.font = reg_font; k.number_format = "0.00"
    ws.cell(row=idx, column=12, value="%").font = reg_font
    m = ws.cell(row=idx, column=13, value=f"=J{idx}*(1+K{idx}/100)"); m.font = reg_font; m.number_format = "$#,##0.00"
    n = ws.cell(row=idx, column=14, value=f"=IF(M{idx}=0,0,O{idx}/M{idx})"); n.font = reg_font; n.number_format = "0.00%"
    o = ws.cell(row=idx, column=15, value=f"=M{idx}-J{idx}"); o.font = reg_font; o.number_format = "$#,##0.00"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Data rows: {len(rows)}")
