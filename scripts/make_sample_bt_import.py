"""Generate sample_BT_import.xlsx matching Buildertrend's import schema."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\sample_BT_import.xlsx"

HEADERS = [
    "Category", "Cost Code", "Title", "Parent Group", "Parent Group Description",
    "Subgroup", "Subgroup Description", "Option Type", "Line Item Type", "Description",
    "Quantity", "Unit", "Unit Cost", "Cost Type", "Marked As", "Builder Cost",
    "Markup", "Markup Type", "Unit Price", "Client Price", "Margin", "Profit",
    "% Invoiced", "Internal Notes",
]

COL_WIDTHS = {
    "A": 34, "B": 42, "C": 32, "D": 34, "E": 22, "F": 16, "G": 22, "H": 12,
    "I": 14, "J": 20, "K": 10, "L": 8, "M": 12, "N": 12, "O": 12, "P": 14,
    "Q": 10, "R": 12, "S": 12, "T": 14, "U": 10, "V": 10, "W": 12, "X": 20,
}

# Item definition: (Category, Title, labor_spec, material_spec)
# Each spec: (cost_code, qty, unit, unit_cost) or None
DEMO_CAT = "02 Existing Conditions"
CARP_CAT = "06 Woods, Plastics, and Composites"
FURN_CAT = "12 Furnishings"
FIN_CAT = "09 Finishes"
PLUMB_CAT = "22 Plumbing"
ELEC_CAT = "26 Electrical"
APP_CAT = "11 Equipment"
GC_CAT = "01 General Requirements"
PERMIT_CAT = "00 Procurement and Contracting Requirements"

DEMO_L = "2.201 Demolition and Removal Labor"
DEMO_M = "2.202 Demolition and Removal Material"
CARP_L = "6.101 Rough Carpentry Labor"
CARP_M = "6.102 Rough Carpentry Material"
CASE_L = "12.101 Casework and Millwork Labor"
CASE_M = "12.102 Casework and Millwork Material"
TILE_L = "9.301 Tile and Stone Labor"
TILE_M = "9.302 Tile and Stone Material"
LVP_L = "9.501 Resilient Flooring Labor"
LVP_M = "9.502 Resilient Flooring Material"
PAINT_L = "9.801 Painting and Coating Labor"
PAINT_M = "9.802 Painting and Coating Material"
PLUMB_L = "22.201 Plumbing Fixtures Labor"
PLUMB_M = "22.202 Plumbing Fixtures Material"
LIGHT_L = "26.301 Lighting Systems Labor"
LIGHT_M = "26.302 Lighting Systems Material"
WIRE_L = "26.201 Wiring and Cabling Labor"
WIRE_M = "26.202 Wiring and Cabling Material"
APP_L = "11.201 Residential and Hospitality Fixtures Labor"
APP_M = "11.202 Residential and Hospitality Fixtures Material"
PM_L = "1.201 Project Personnel Costs Labor"
CLEAN_L = "1.901 Cleaning and Site Maintenance Labor"
CLEAN_M = "1.902 Cleaning and Site Maintenance Material"
INS_L = "1.701 Bonding and Insurance Labor"
PERMIT_L = "0.101 Permits and Approvals Labor"

ITEMS = [
    # Demolition
    (DEMO_CAT, "Cabinet Demolition",
        (DEMO_L, 8, "HR", 85), (DEMO_M, 22, "LF", 25)),
    (DEMO_CAT, "Countertop & Backsplash Removal",
        (DEMO_L, 4, "HR", 85), (DEMO_M, 38, "SF", 8)),
    (DEMO_CAT, "Load-Bearing Wall Removal w/ Beam",
        (DEMO_L, 16, "HR", 100), (DEMO_M, 1, "EA", 1200)),
    # Rough Carpentry
    (CARP_CAT, "Blocking and Backing",
        (CARP_L, 6, "HR", 85), (CARP_M, 1, "LS", 280)),
    (CARP_CAT, "Structural Beam Installation",
        (CARP_L, 12, "HR", 100), (CARP_M, 1, "EA", 1800)),
    # Cabinetry
    (FURN_CAT, "White Shaker Cabinets - Perimeter",
        (CASE_L, 32, "HR", 85), (CASE_M, 22, "LF", 400)),
    (FURN_CAT, "Island Base Cabinets",
        (CASE_L, 12, "HR", 85), (CASE_M, 8, "LF", 400)),
    (FURN_CAT, "Soft-Close Hardware Package",
        (CASE_L, 4, "HR", 85), (CASE_M, 1, "LOT", 720)),
    # Countertops
    (FURN_CAT, "Quartz Countertop - Perimeter",
        (CASE_L, 6, "HR", 85), (CASE_M, 38, "SF", 120)),
    (FURN_CAT, "Waterfall Edge Island Top",
        (CASE_L, 8, "HR", 85), (CASE_M, 24, "SF", 180)),
    # Finishes
    (FIN_CAT, "Subway Tile Backsplash",
        (TILE_L, 16, "HR", 85), (TILE_M, 38, "SF", 18)),
    (FIN_CAT, "LVP Flooring",
        (LVP_L, 12, "HR", 85), (LVP_M, 280, "SF", 8)),
    (FIN_CAT, "Paint Walls and Ceiling",
        (PAINT_L, 18, "HR", 85), (PAINT_M, 1, "LS", 380)),
    # Plumbing
    (PLUMB_CAT, "Sink & Faucet Rough-in and Install",
        (PLUMB_L, 6, "HR", 130), (PLUMB_M, 1, "LS", 850)),
    # Electrical
    (ELEC_CAT, 'Recessed Can Lights (6")',
        (LIGHT_L, 8, "HR", 85), (LIGHT_M, 8, "EA", 85)),
    (ELEC_CAT, "Under-Cabinet LED Lighting",
        (LIGHT_L, 6, "HR", 85), (LIGHT_M, 22, "LF", 28)),
    (ELEC_CAT, "Pendant Lights Above Island",
        (LIGHT_L, 3, "HR", 85), (LIGHT_M, 3, "EA", 220)),
    (ELEC_CAT, "Outlets & Switch Relocation",
        (WIRE_L, 8, "HR", 85), (WIRE_M, 6, "EA", 45)),
    # Appliances
    (APP_CAT, "Stainless Appliance Package",
        (APP_L, 4, "HR", 85), (APP_M, 1, "LOT", 9500)),
    # General Conditions
    (GC_CAT, "Project Management",
        (PM_L, 40, "HR", 130), None),
    (GC_CAT, "Dumpster & Disposal",
        (CLEAN_L, 2, "HR", 85), (CLEAN_M, 1, "LS", 650)),
    (GC_CAT, "General Liability Insurance",
        (INS_L, 1, "LS", 1100), None),
    # Permits
    (PERMIT_CAT, "Building Permit",
        (PERMIT_L, 1, "LS", 650), None),
]


def build_rows():
    rows = []
    for category, title, labor, material in ITEMS:
        parent_group = category
        if labor:
            code, qty, unit, uc = labor
            rows.append({
                "Category": category, "Cost Code": code, "Title": title,
                "Parent Group": parent_group,
                "Quantity": qty, "Unit": unit, "Unit Cost": uc,
                "Cost Type": "Labor",
            })
        if material:
            code, qty, unit, uc = material
            rows.append({
                "Category": category, "Cost Code": code, "Title": title,
                "Parent Group": parent_group,
                "Quantity": qty, "Unit": unit, "Unit Cost": uc,
                "Cost Type": "Material",
            })
    return rows


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Estimate"

    # Styles
    arial = Font(name="Arial", size=10)
    arial_bold_white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")
    zebra_fill = PatternFill("solid", fgColor="F5F5F5")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Column widths
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Header row
    ws.row_dimensions[1].height = 30
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = arial_bold_white
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    rows = build_rows()
    currency_fmt = '"$"#,##0.00'
    num_fmt = '#,##0.00'
    pct_cols = {11, 17, 23}  # K Quantity, Q Markup, W % Invoiced -> numeric fmt
    currency_cols = {13, 16, 19, 20, 21, 22}  # M, P, S, T, U, V

    for idx, r in enumerate(rows):
        excel_row = idx + 2
        is_zebra = (excel_row % 2) == 1  # rows 3, 5, 7, ...
        vals = {
            1: r["Category"],
            2: r["Cost Code"],
            3: r["Title"],
            4: r["Parent Group"],
            5: "",  # Parent Group Description
            6: "",  # Subgroup
            7: "",  # Subgroup Description
            8: "",  # Option Type
            9: "Estimate",  # Line Item Type
            10: "",  # Description
            11: r["Quantity"],
            12: r["Unit"],
            13: r["Unit Cost"],
            14: r["Cost Type"],
            15: "",  # Marked As
            16: f"=K{excel_row}*M{excel_row}",  # Builder Cost
            17: 20,  # Markup
            18: "%",
            19: f"=M{excel_row}*(1+Q{excel_row}/100)",  # Unit Price
            20: f"=P{excel_row}*(1+Q{excel_row}/100)",  # Client Price
            21: f"=T{excel_row}-P{excel_row}",  # Margin
            22: f"=T{excel_row}-P{excel_row}",  # Profit
            23: 0,  # % Invoiced
            24: "",  # Internal Notes
        }
        for col_idx, v in vals.items():
            c = ws.cell(row=excel_row, column=col_idx, value=v)
            c.font = arial
            c.border = border
            if is_zebra:
                c.fill = zebra_fill
            if col_idx in currency_cols:
                c.number_format = currency_fmt
                c.alignment = right_align
            elif col_idx in pct_cols:
                c.number_format = num_fmt
                c.alignment = right_align
            else:
                c.alignment = left_align

    # Freeze panes
    ws.freeze_panes = "A2"

    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Data rows: {len(rows)}")


if __name__ == "__main__":
    main()
