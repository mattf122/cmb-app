"""Final audit: proper proposal table parsing + comprehensive CSI check."""
import re
from openpyxl import load_workbook

ESTIMATE = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Estimate_2026-04-24.xlsx"
PROPOSAL = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Proposal_2026-04-24.doc"
CSI_REF = r"C:\Users\MattFarrier\Downloads\4.17.26Buildertrend_CSI_Cost_Codes.xlsx"


def money(s):
    return float(s.replace("$","").replace(",","").strip())


# Load CSI
wb = load_workbook(CSI_REF, data_only=True)
ws = wb.active
csi_codes = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or row[0] is None:
        continue
    code = str(row[0]).strip()
    div = str(row[1]).strip() if len(row) > 1 and row[1] else ""
    csi_codes[code] = div

print(f"CSI codes loaded: {len(csi_codes)}")
# Sample codes format
print("Sample CSI codes (first 10):")
for c in list(csi_codes)[:10]:
    print(f"  {c!r}")
# What division format? e.g. "01 General Requirements" - check how division prefix looks
print("\nDivisions sample:")
divs = sorted(set(csi_codes.values()))
for d in divs[:10]:
    print(f"  {d!r}")

# Now look at what CSI code prefixes look like
prefixes = set()
for c in csi_codes:
    m = re.match(r"(\d+\.\d+)", c)
    if m:
        prefixes.add(m.group(1))
# Check if "01.101" or "1.101" format
for c in list(csi_codes)[:3]:
    print(f"Format test: {c}")

# Parse Excel
wb2 = load_workbook(ESTIMATE, data_only=True)
ws2 = wb2.active
rows = []
for r in range(2, ws2.max_row + 1):
    row = [ws2.cell(row=r, column=c).value for c in range(1, 16)]
    if all(v is None or v == "" for v in row):
        continue
    rows.append((r, row))

# Revalidate: try matching with code's leading token only
print("\n=== Cost code validation (v2) ===")
SUFFIX_TO_TYPE = {"01": "Labor", "02": "Material", "03": "Subcontractor"}
issues = []
unique_codes_in_excel = set()
for r, row in rows:
    cat = row[0]
    code = row[1]
    title = row[2]
    ctype = row[7]
    code_s = str(code or "").strip()
    unique_codes_in_excel.add(code_s)

    problems = []
    # Is the full string in CSI list?
    exact = code_s in csi_codes
    # Extract numeric prefix like "01.101"
    # Try first token
    first_tok = code_s.split()[0] if code_s else ""
    # Check if first token matches any csi key's numeric part
    matching_csi = [c for c in csi_codes if c.startswith(first_tok + " ")]
    if not exact and not matching_csi:
        # Try with leading zero stripped or added
        alt = first_tok.lstrip("0") if first_tok else ""
        matching_csi2 = [c for c in csi_codes if c.startswith(alt + " ")]
        if matching_csi2:
            problems.append(f"leading-zero format issue (code '{first_tok}' should be '{alt}')")
        else:
            # Also try adding leading zero
            alt2 = "0" + first_tok if first_tok and not first_tok.startswith("0") else ""
            matching_csi3 = [c for c in csi_codes if c.startswith(alt2 + " ")]
            if matching_csi3:
                problems.append(f"missing-leading-zero (code '{first_tok}' should be '{alt2}')")
            else:
                problems.append(f"prefix '{first_tok}' not found in CSI list")
    elif not exact and matching_csi:
        problems.append(f"code includes label text; CSI prefix '{first_tok}' valid but full value not a CSI key")

    # suffix
    m = re.match(r"^\d+\.(\d+)", first_tok)
    suffix = ""
    if m:
        digits = m.group(1)
        if len(digits) >= 2:
            suffix = digits[-2:]
    if suffix in SUFFIX_TO_TYPE:
        expected = SUFFIX_TO_TYPE[suffix]
        if ctype != expected:
            problems.append(f"suffix {suffix} expects {expected}, got '{ctype}'")
    else:
        problems.append(f"cannot extract 01/02/03 suffix from '{first_tok}'")

    # division
    code_div_m = re.match(r"^(\d+)\.", first_tok)
    code_div = code_div_m.group(1).lstrip("0") if code_div_m else ""
    cat_div_m = re.match(r"^(\d+)", str(cat or ""))
    cat_div = cat_div_m.group(1).lstrip("0") if cat_div_m else ""
    if code_div and cat_div and code_div != cat_div:
        problems.append(f"code div {code_div} != category div {cat_div}")

    if problems:
        issues.append((r, cat, code, title, ctype, "; ".join(problems)))

print(f"Rows with issues: {len(issues)} of {len(rows)}")

# Break down problem types
print("\nProblem type counts:")
from collections import Counter
probs = Counter()
for iss in issues:
    for p in iss[5].split("; "):
        key = re.sub(r"'[^']*'", "'X'", p)
        probs[key] += 1
for k, v in probs.most_common():
    print(f"  {v:3d}  {k}")

# Unique codes in excel
print(f"\nUnique cost code strings in Excel: {len(unique_codes_in_excel)}")
for c in sorted(unique_codes_in_excel)[:25]:
    print(f"  {c!r}")

# Totals computed manually (col J and M are formulas)
total_bc = 0.0
total_cp = 0.0
per_cat = {}
for r, row in rows:
    qty = row[4] or 0
    uc = row[6] or 0
    markup = row[10] or 0
    cat = row[0]
    try:
        bc = float(qty) * float(uc)
    except Exception:
        bc = 0
    try:
        cp = bc * (1 + float(markup)/100)
    except Exception:
        cp = bc
    total_bc += bc
    total_cp += cp
    per_cat.setdefault(cat, {"bc":0.0,"cp":0.0,"n":0})
    per_cat[cat]["bc"] += bc
    per_cat[cat]["cp"] += cp
    per_cat[cat]["n"] += 1

print(f"\n=== TOTALS ===")
print(f"Excel Builder Cost total: ${total_bc:,.2f}")
print(f"Excel Client Price total: ${total_cp:,.2f}")

# Parse proposal HTML table structure
with open(PROPOSAL, "rb") as f:
    html = f.read().decode("utf-8", errors="replace")

# Find all table rows with dollar cells
# Each section row pattern: <tr><td>Name</td><td ...>Division</td><td ...>$low</td><td ...>$high</td></tr>
tr_pat = re.compile(r"<tr[^>]*>(.*?)</tr>", re.DOTALL)
proposal_rows = []
for tr in tr_pat.findall(html):
    # Get all td contents
    tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.DOTALL)
    # Strip tags
    tds_text = [re.sub(r"<[^>]+>", "", t).strip() for t in tds]
    if not tds_text:
        continue
    # Find dollars
    dollars = [t for t in tds_text if t.startswith("$")]
    if len(dollars) >= 2:
        # pick first 2 dollar cells as low/high
        try:
            lo = money(dollars[0])
            hi = money(dollars[1])
        except Exception:
            continue
        # Name is first non-dollar cell
        label_cells = [t for t in tds_text if not t.startswith("$")]
        name = label_cells[0] if label_cells else ""
        div = label_cells[1] if len(label_cells) > 1 else ""
        proposal_rows.append((name, div, lo, hi))

print(f"\nProposal table rows with low/high: {len(proposal_rows)}")
for pr in proposal_rows:
    print(f"  {pr[0]:<35} | {pr[1]:<30} | ${pr[2]:>10,.2f} - ${pr[3]:>10,.2f}")

# Identify grand total: the row marked TOTAL PROJECT COST
grand_lo = grand_hi = None
for pr in proposal_rows:
    if "TOTAL PROJECT COST" in pr[0]:
        grand_lo, grand_hi = pr[2], pr[3]
        break
if grand_lo is None:
    # fallback to last row
    grand_lo, grand_hi = proposal_rows[-1][2], proposal_rows[-1][3]

print(f"\nProposal GRAND TOTAL: ${grand_lo:,.2f} – ${grand_hi:,.2f}")

# Construction subtotal
cs_lo = cs_hi = None
for pr in proposal_rows:
    if "Construction Subtotal" in pr[0]:
        cs_lo, cs_hi = pr[2], pr[3]
        break
if cs_lo:
    print(f"Proposal Construction Subtotal: ${cs_lo:,.2f} – ${cs_hi:,.2f}")

# Reconcile
print("\n=== RECONCILIATION ===")
print(f"Excel Builder Cost  ${total_bc:>12,.2f}")
print(f"Excel Client Price  ${total_cp:>12,.2f}")
print(f"Doc Grand Total Lo  ${grand_lo:>12,.2f}")
print(f"Doc Grand Total Hi  ${grand_hi:>12,.2f}")
if grand_hi:
    pct_b = (total_bc - grand_hi) / grand_hi * 100
    pct_c = (total_cp - grand_hi) / grand_hi * 100
    print(f"Excel BC vs doc HIGH: {pct_b:+.1f}%")
    print(f"Excel CP vs doc HIGH: {pct_c:+.1f}%")
if grand_lo:
    pct_b = (total_bc - grand_lo) / grand_lo * 100
    pct_c = (total_cp - grand_lo) / grand_lo * 100
    print(f"Excel BC vs doc LOW:  {pct_b:+.1f}%")
    print(f"Excel CP vs doc LOW:  {pct_c:+.1f}%")

# Per-section: match category name in Excel to rows in proposal
# Proposal rows have "Name" (trade name like "Plumbing") and "Division" (like "22 Plumbing")
# Aggregate proposal by division
prop_by_div = {}
for name, div, lo, hi in proposal_rows:
    if not div or "Subtotal" in name or "TOTAL" in name.upper() or "General Conditions" == name and "General" in div:
        pass
    if div and div != "" and div[0].isdigit():
        prop_by_div.setdefault(div, {"lo":0.0,"hi":0.0,"names":[]})
        prop_by_div[div]["lo"] += lo
        prop_by_div[div]["hi"] += hi
        prop_by_div[div]["names"].append(name)

print("\nProposal aggregated by division:")
for d, v in sorted(prop_by_div.items()):
    print(f"  {d:<30} lo=${v['lo']:>10,.2f} hi=${v['hi']:>10,.2f} names={v['names']}")

print("\nPer-category side-by-side:")
print(f"  {'Category':<30} {'n':<4} {'Excel BC':>12} {'Doc Low':>12} {'Doc High':>12} {'Flag':<20}")
excel_cats = set(str(c) for c in per_cat)
doc_cats = set(prop_by_div)
all_cats = sorted(excel_cats | doc_cats)
for cat in all_cats:
    ev = per_cat.get(cat, {"bc":0,"cp":0,"n":0})
    pv = prop_by_div.get(cat, {"lo":0,"hi":0})
    flag = ""
    if ev["bc"] > 0 and pv["hi"] > 0:
        if ev["bc"] > pv["hi"] * 1.15:
            flag = f">+15% of high ({(ev['bc']/pv['hi']-1)*100:+.0f}%)"
        elif ev["bc"] < pv["lo"] * 0.85:
            flag = f"<-15% of low ({(ev['bc']/pv['lo']-1)*100:+.0f}%)"
    elif ev["bc"] > 0 and pv["hi"] == 0:
        flag = "in Excel not in Doc"
    elif pv["hi"] > 0 and ev["bc"] == 0:
        flag = "in Doc not in Excel"
    print(f"  {cat:<30} {ev['n']:<4} ${ev['bc']:>11,.0f} ${pv['lo']:>11,.0f} ${pv['hi']:>11,.0f}  {flag}")

# First 10 bad rows
print("\n=== First 10 bad rows ===")
for iss in issues[:10]:
    r, cat, code, title, ct, prob = iss
    print(f"  Row {r}: Cat='{cat}' | Code='{code}' | Title='{title}' | Type='{ct}'")
    print(f"    -> {prob}")
