"""Quick summary: all row values in compact form."""
from openpyxl import load_workbook
TEMPLATE = r"C:\Users\MattFarrier\Downloads\New Construction Proposal - 2000 sqft Custom Home v2.xlsx"
wb = load_workbook(TEMPLATE, data_only=False)
ws = wb['Sheet1']

# Column headers
hdr = [ws.cell(1, c).value for c in range(1, 16)]
print("HEADERS:", hdr)

# Unique values per categorical col
cats = set()
cost_types = set()
units = set()
markup_types = set()
marked_as = set()
for r in range(2, ws.max_row+1):
    cats.add(ws.cell(r, 1).value)
    units.add(ws.cell(r, 6).value)
    cost_types.add(ws.cell(r, 8).value)
    marked_as.add(ws.cell(r, 9).value)
    markup_types.add(ws.cell(r, 12).value)

print("\nUnique Categories (col A):")
for v in sorted(str(x) for x in cats if x is not None):
    print(" ", v)
print("\nUnique Units (col F):", sorted(str(x) for x in units if x is not None))
print("Unique Cost Types (col H):", sorted(str(x) for x in cost_types if x is not None))
print("Unique Marked As (col I):", sorted(str(x) for x in marked_as if x is not None))
print("Unique Markup Types (col L):", sorted(str(x) for x in markup_types if x is not None))

# Print every row in compact form
print("\nALL ROWS (compact):")
for r in range(1, ws.max_row+1):
    vals = [ws.cell(r, c).value for c in range(1, 16)]
    print(f"r{r}:", vals)
