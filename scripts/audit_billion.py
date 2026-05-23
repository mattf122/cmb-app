"""Audit the Billion estimate Excel + proposal doc against CSI codes."""
import re
from pathlib import Path
from openpyxl import load_workbook

ESTIMATE = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Estimate_2026-04-24.xlsx"
PROPOSAL = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\2026\Billion - 1999\Billion_Proposal_2026-04-24.doc"
CSI_REF = r"C:\Users\MattFarrier\Downloads\4.17.26Buildertrend_CSI_Cost_Codes.xlsx"

EXPECTED_HEADERS = [
    "Category", "Cost Code", "Title", "Description", "Quantity", "Unit",
    "Unit Cost", "Cost Type", "Marked As", "Builder Cost", "Markup",
    "Markup Type", "Client Price", "Margin", "Profit"
]

SUFFIX_TO_TYPE = {"01": "Labor", "02": "Material", "03": "Subcontractor"}


def load_csi_codes():
    wb = load_workbook(CSI_REF, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    print(f"CSI ref headers: {header}")
    codes = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        code = str(r[0]).strip()
        division = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        codes[code] = division
    return codes


def audit_excel(csi_codes):
    wb = load_workbook(ESTIMATE, data_only=True)
    ws = wb.active
    print(f"\n=== SECTION A: Excel structure ===")
    print(f"Sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

    row1 = [ws.cell(row=1, column=c).value for c in range(1, 16)]
    print(f"Row 1: {row1}")
    headers_match = row1 == EXPECTED_HEADERS
    print(f"Headers match expected 15: {headers_match}")
    if not headers_match:
        for i, (got, want) in enumerate(zip(row1, EXPECTED_HEADERS), 1):
            if got != want:
                print(f"  col {i}: got '{got}' want '{want}'")

    # Check for title/client rows by inspecting row 2 to see if it's data
    row2 = [ws.cell(row=2, column=c).value for c in range(1, 16)]
    print(f"Row 2 (first data row): {row2}")

    data = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, 16)]
        if all(v is None or v == "" for v in row):
            continue
        data.append((r, row))
    print(f"Data rows: {len(data)}")

    print(f"\n=== SECTION B: Cost code validation ===")
    issues = []
    for r, row in data:
        category = row[0]
        code = row[1]
        title = row[2]
        cost_type = row[7]
        problems = []

        code_str = str(code).strip() if code is not None else ""
        if not code_str:
            problems.append("empty cost code")
        elif code_str not in csi_codes:
            problems.append("not in CSI list")

        # suffix check
        suffix = code_str.split("-")[-1] if "-" in code_str else code_str[-2:]
        # try also last 2 chars after dot
        # Typical CSI format: "3.1-01" or similar. Extract final digits group
        m = re.search(r"(\d{2})\s*$", code_str)
        suffix_found = m.group(1) if m else ""
        expected_type = SUFFIX_TO_TYPE.get(suffix_found)
        if expected_type is None:
            problems.append(f"suffix '{suffix_found}' not 01/02/03")
        elif cost_type != expected_type:
            problems.append(f"suffix {suffix_found} should be {expected_type}, got '{cost_type}'")

        # division check - prefix before dot in code
        code_div_m = re.match(r"(\d+)\.", code_str)
        code_div = code_div_m.group(1) if code_div_m else ""
        cat_div_m = re.match(r"(\d+)\.", str(category or ""))
        cat_div = cat_div_m.group(1) if cat_div_m else ""
        if code_div and cat_div and code_div != cat_div:
            problems.append(f"code div {code_div} != category div {cat_div}")

        if problems:
            issues.append((r, category, code, title, cost_type, "; ".join(problems)))

    print(f"Total issues: {len(issues)} of {len(data)} rows")

    # Print first 50 issues as table
    print("\nFirst 50 issue rows:")
    print(f"{'Row':<5} {'Category':<35} {'Code':<12} {'Title':<35} {'Type':<14} {'Problem'}")
    for iss in issues[:50]:
        r, cat, code, title, ct, prob = iss
        cat_s = str(cat or "")[:34]
        code_s = str(code or "")[:11]
        title_s = str(title or "")[:34]
        ct_s = str(ct or "")[:13]
        print(f"{r:<5} {cat_s:<35} {code_s:<12} {title_s:<35} {ct_s:<14} {prob}")

    # Totals
    print(f"\n=== SECTION C: Totals ===")
    builder_total = 0.0
    client_total = 0.0
    per_cat = {}
    for r, row in data:
        bc = row[9]
        cp = row[12]
        cat = row[0]
        try:
            bc_v = float(bc) if bc is not None else 0.0
        except (TypeError, ValueError):
            bc_v = 0.0
        try:
            cp_v = float(cp) if cp is not None else 0.0
        except (TypeError, ValueError):
            cp_v = 0.0
        builder_total += bc_v
        client_total += cp_v
        per_cat.setdefault(cat, {"bc": 0.0, "cp": 0.0, "n": 0})
        per_cat[cat]["bc"] += bc_v
        per_cat[cat]["cp"] += cp_v
        per_cat[cat]["n"] += 1

    print(f"Builder Cost total (col J): ${builder_total:,.2f}")
    print(f"Client Price total  (col M): ${client_total:,.2f}")

    return data, issues, builder_total, client_total, per_cat


def parse_proposal():
    print(f"\n=== Parsing proposal HTML ===")
    with open(PROPOSAL, "rb") as f:
        raw = f.read()
    # Try a few decodings
    for enc in ("utf-8", "utf-16", "latin-1"):
        try:
            text = raw.decode(enc)
            print(f"Decoded with {enc}, len={len(text)}")
            break
        except UnicodeDecodeError:
            continue

    # Strip HTML tags to get plain text version
    plain = re.sub(r"<[^>]+>", " ", text)
    plain = re.sub(r"\s+", " ", plain)
    plain = re.sub(r"&nbsp;", " ", plain)
    plain = re.sub(r"&amp;", "&", plain)

    # Find all dollar ranges: $X,XXX – $Y,YYY or with - or en-dash
    # Range patterns
    range_pat = re.compile(r"\$([\d,]+(?:\.\d+)?)\s*[–\-—]\s*\$([\d,]+(?:\.\d+)?)")
    ranges = range_pat.findall(plain)
    print(f"Found {len(ranges)} dollar ranges in proposal")

    # Try to find section names near each range
    # Look for text segments with header-like content
    # Print a chunk of plain text to eyeball
    # Find grand total or bottom line
    grand = None
    for m in range(len(ranges) - 1, -1, -1):
        lo, hi = ranges[m]
        # last range likely grand total
        grand = (lo, hi)
        break

    # Try to extract per-section labels. Look for repeated pattern: Section Name ... $low – $high
    # Use HTML structure with tags
    # Search for <td> or <tr> cells
    sections = []
    # A common pattern in the generator: each row has name then range
    # Try regex across text that captures a word phrase before a range
    section_pat = re.compile(
        r"([A-Z][A-Za-z&/\- ]{2,60}?)\s*\$([\d,]+(?:\.\d+)?)\s*[–\-—]\s*\$([\d,]+(?:\.\d+)?)"
    )
    for m in section_pat.finditer(plain):
        name, lo, hi = m.groups()
        sections.append((name.strip(), lo, hi))
    print(f"Section-range matches: {len(sections)}")
    for s in sections[:40]:
        print(f"  {s}")

    return text, plain, ranges, sections, grand


def main():
    csi = load_csi_codes()
    print(f"Loaded {len(csi)} CSI codes")
    # Show a few samples
    sample_items = list(csi.items())[:5]
    for c, d in sample_items:
        print(f"  {c!r} -> {d!r}")

    data, issues, b_total, c_total, per_cat = audit_excel(csi)

    text, plain, ranges, sections, grand = parse_proposal()

    print(f"\n=== SECTION C (cont): Reconciliation ===")
    if grand:
        glo = float(grand[0].replace(",", ""))
        ghi = float(grand[1].replace(",", ""))
        print(f"Proposal grand range (last found): ${glo:,.2f} – ${ghi:,.2f}")
        for label, total in [("Builder Cost", b_total), ("Client Price", c_total)]:
            diff_hi = (total - ghi) / ghi * 100 if ghi else 0
            diff_lo = (total - glo) / glo * 100 if glo else 0
            print(f"  Excel {label} ${total:,.2f}: {diff_hi:+.1f}% vs doc HIGH, {diff_lo:+.1f}% vs doc LOW")

    print(f"\n=== SECTION D: Per-category reconciliation ===")
    for cat, v in sorted(per_cat.items(), key=lambda x: str(x[0])):
        print(f"  {str(cat or '<blank>'):<45} n={v['n']:<4} BC=${v['bc']:>12,.2f}  CP=${v['cp']:>12,.2f}")

    print(f"\n=== SECTION E: First 10 bad rows ===")
    for iss in issues[:10]:
        r, cat, code, title, ct, prob = iss
        print(f"  Row {r}: Cat='{cat}' | Code='{code}' | Title='{title}' | Type='{ct}' -> {prob}")

    print(f"\n=== Summary ===")
    print(f"Data rows: {len(data)}")
    print(f"Invalid/mismatched cost code rows: {len(issues)}")
    print(f"Builder total: ${b_total:,.2f}")
    print(f"Client total: ${c_total:,.2f}")


if __name__ == "__main__":
    main()
