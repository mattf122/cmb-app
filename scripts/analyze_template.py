"""Exhaustive analysis of a proposal Excel template."""
import sys
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OUT_PATH = r"C:\Users\MattFarrier\OneDrive - Copper Mountain Builders\CMB Site Visits\App Code\scripts\template_analysis.txt"
TEMPLATE = r"C:\Users\MattFarrier\Downloads\New Construction Proposal - 2000 sqft Custom Home v2.xlsx"
COSTCODES = r"C:\Users\MattFarrier\Downloads\4.17.26Buildertrend_CSI_Cost_Codes.xlsx"

out_lines = []
def p(*args):
    s = " ".join(str(a) for a in args)
    out_lines.append(s)
    print(s)

def color_hex(c):
    if c is None:
        return None
    try:
        if c.type == 'rgb' and c.rgb:
            return c.rgb
        if c.type == 'theme':
            return f"theme={c.theme},tint={c.tint}"
        if c.type == 'indexed':
            return f"indexed={c.value}"
    except Exception:
        pass
    return str(getattr(c, 'rgb', c))

def describe_font(f):
    if not f:
        return "None"
    return f"name={f.name} size={f.size} bold={f.bold} italic={f.italic} color={color_hex(f.color)}"

def describe_fill(fill):
    if not fill:
        return "None"
    try:
        pt = fill.patternType
        if not pt:
            return "no-fill"
        fg = color_hex(fill.fgColor)
        bg = color_hex(fill.bgColor)
        return f"pattern={pt} fg={fg} bg={bg}"
    except Exception as e:
        return f"err:{e}"

def describe_border(b):
    if not b:
        return "None"
    parts = []
    for side in ['left','right','top','bottom']:
        s = getattr(b, side)
        if s and s.style:
            parts.append(f"{side}={s.style}/{color_hex(s.color)}")
    return ", ".join(parts) if parts else "none"

def describe_alignment(a):
    if not a:
        return ""
    return f"h={a.horizontal} v={a.vertical} wrap={a.wrap_text} indent={a.indent}"

def analyze_sheet(ws, max_data_rows=80):
    p("\n" + "="*100)
    p(f"SHEET: {ws.title!r}")
    p("="*100)
    p(f"Dimensions: {ws.dimensions}  (max_row={ws.max_row}, max_col={ws.max_column})")
    p(f"Sheet state: {ws.sheet_state}")

    # Freeze
    p(f"Frozen panes: {ws.freeze_panes}")

    # Merged cells
    merges = sorted(str(r) for r in ws.merged_cells.ranges)
    p(f"Merged cells ({len(merges)}):")
    for m in merges:
        p(f"  {m}")

    # Column widths
    p("Column widths:")
    for col_letter, dim in sorted(ws.column_dimensions.items()):
        if dim.width is not None:
            p(f"  {col_letter}: width={dim.width} hidden={dim.hidden}")

    # Row heights (only non-default)
    p("Row heights (non-default):")
    for row_idx, dim in sorted(ws.row_dimensions.items()):
        if dim.height is not None:
            p(f"  row {row_idx}: height={dim.height}")

    # Images
    try:
        imgs = ws._images
        p(f"Images: {len(imgs)}")
        for i, im in enumerate(imgs):
            anc = getattr(im, 'anchor', None)
            p(f"  image {i}: anchor={anc}")
    except Exception as e:
        p(f"Images error: {e}")

    # Charts
    try:
        p(f"Charts: {len(ws._charts)}")
    except Exception:
        pass

    # Print every cell with content + formatting
    p("\n--- CELL-BY-CELL DUMP (rows with content) ---")
    rows_dumped = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        row_has_content = any(c.value is not None for c in row)
        if not row_has_content:
            continue
        rows_dumped += 1
        if rows_dumped > max_data_rows:
            p(f"  ... truncated after {max_data_rows} non-empty rows (total max_row={ws.max_row})")
            break
        r = row[0].row
        p(f"\nRow {r}:")
        for c in row:
            if c.value is None and (c.fill is None or c.fill.patternType is None) and c.font.name is None:
                continue
            val = c.value
            if isinstance(val, str) and len(val) > 200:
                val = val[:200] + "..."
            p(f"  {c.coordinate}: value={val!r}")
            p(f"    number_format={c.number_format!r}")
            p(f"    font: {describe_font(c.font)}")
            p(f"    fill: {describe_fill(c.fill)}")
            p(f"    border: {describe_border(c.border)}")
            p(f"    align: {describe_alignment(c.alignment)}")
            if isinstance(c.value, str) and c.value.startswith('='):
                p(f"    FORMULA: {c.value}")

# ============ TEMPLATE ============
p("#"*100)
p(f"FILE: {TEMPLATE}")
p("#"*100)
wb = load_workbook(TEMPLATE, data_only=False)
p(f"Sheet names: {wb.sheetnames}")
p(f"Defined names: {list(wb.defined_names)}")
for name in wb.sheetnames:
    analyze_sheet(wb[name], max_data_rows=200)

# Also load with data_only to see computed values for formulas
p("\n" + "#"*100)
p("FORMULAS WITH COMPUTED VALUES (data_only=True)")
p("#"*100)
wb2 = load_workbook(TEMPLATE, data_only=True)
wb_f = load_workbook(TEMPLATE, data_only=False)
for name in wb_f.sheetnames:
    ws_f = wb_f[name]
    ws_v = wb2[name]
    p(f"\n-- Sheet {name} formulas --")
    count = 0
    for row in ws_f.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                computed = ws_v[c.coordinate].value
                p(f"  {c.coordinate}: {c.value}  =>  {computed!r}")
                count += 1
                if count > 60:
                    break
        if count > 60:
            p("  ... truncated")
            break

# ============ COST CODES ============
p("\n" + "#"*100)
p(f"FILE: {COSTCODES}")
p("#"*100)
wb3 = load_workbook(COSTCODES, data_only=True)
p(f"Sheet names: {wb3.sheetnames}")
for name in wb3.sheetnames:
    ws = wb3[name]
    p(f"\nSheet {name}: {ws.dimensions} rows={ws.max_row} cols={ws.max_column}")
    # headers
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    p(f"Headers row 1: {headers}")
    # first 10 and last 5 rows
    p("First 10 data rows:")
    for r in range(2, min(12, ws.max_row+1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        p(f"  row {r}: {vals}")
    p("Last 5 rows:")
    for r in range(max(2, ws.max_row-4), ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        p(f"  row {r}: {vals}")

with open(OUT_PATH, 'w', encoding='utf-8') as f:
    f.write("\n".join(out_lines))
print(f"\nWrote {OUT_PATH} ({len(out_lines)} lines)")
